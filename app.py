import io
import hashlib
from datetime import date, timedelta

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

st.set_page_config(page_title="Aurora · Precios y Promos", layout="wide")

# ---------- Styling ----------
st.markdown(
    """
    <style>
    .small-note {color:#6b7280; font-size:0.88rem;}
    .chip {display:inline-block; padding:0.22rem 0.55rem; border:1px solid #d1d5db; border-radius:999px; margin:0.15rem 0.25rem 0.15rem 0; font-size:0.85rem; background:#f9fafb;}
    .section-card {border:1px solid #e5e7eb; border-radius:16px; padding:0.9rem 1rem; background:white;}
    .promo-card {border:1px solid #e5e7eb; border-radius:16px; padding:0.7rem 0.85rem; background:white; min-height:122px;}
    .status-dot {font-size:0.85rem; font-weight:600;}
    .muted {color:#6b7280;}
    div.stButton > button[kind="secondary"] {
        border-radius: 18px;
        border: 1px solid #d1d5db;
        background: white;
        padding: 0.85rem 0.9rem;
        min-height: 118px;
        text-align: left;
        white-space: pre-line;
        line-height: 1.25;
        width: 100%;
    }
    div.stButton > button[kind="secondary"]:hover {
        border-color: #9ca3af;
        background: #f9fafb;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# ---------- Helpers ----------
TODAY = date.today()
VAT = 1.19


def file_hash(uploaded_file) -> str:
    if uploaded_file is None:
        return ""
    return hashlib.md5(uploaded_file.getvalue()).hexdigest()


def clean_sku(value) -> str:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return ""
    s = str(value).strip()
    if not s or s.lower() == "nan":
        return ""
    if s.endswith(".0"):
        s = s[:-2]
    try:
        if "." in s:
            f = float(s)
            if f.is_integer():
                s = str(int(f))
    except Exception:
        pass
    return s.replace(" ", "")


def clean_text(value) -> str:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return ""
    s = str(value).strip()
    return "" if s.lower() == "nan" else s


def to_float(value):
    if value is None:
        return np.nan
    if isinstance(value, str):
        value = value.strip().replace("$", "").replace(".", "") if value.count(",") == 1 and value.count(".") > 1 else value.strip()
        value = value.replace(",", ".")
    try:
        return float(value)
    except Exception:
        return np.nan


def parse_date(value):
    if value is None or value == "":
        return pd.NaT
    if isinstance(value, pd.Timestamp):
        return value.normalize()
    dt = pd.to_datetime(value, errors="coerce", dayfirst=True)
    if pd.isna(dt):
        return pd.NaT
    return pd.Timestamp(dt.date())


def fmt_date(value) -> str:
    dt = parse_date(value)
    if pd.isna(dt):
        return "—"
    return dt.strftime("%d/%m/%Y")


def status_from_date(value) -> str:
    dt = parse_date(value)
    if pd.isna(dt):
        return "Sin fecha"
    d = dt.date()
    delta = (d - TODAY).days
    if delta < 0:
        return "Vencida"
    if delta == 0:
        return "Vence hoy"
    if delta == 1:
        return "Vence mañana"
    if delta == 3:
        return "Vence en 3 días"
    if delta <= 7:
        return "Próx. 7 días"
    return "Vigente"


def margin_local(cost, bruto):
    cost = to_float(cost)
    bruto = to_float(bruto)
    if np.isnan(cost) or np.isnan(bruto) or bruto <= 0:
        return np.nan, np.nan
    neto = bruto / VAT
    if neto == 0:
        return np.nan, np.nan
    margin = (neto - cost) / neto
    return neto, margin


def margin_meli1(cost, monto_sim):
    cost = to_float(cost)
    monto_sim = to_float(monto_sim)
    if np.isnan(cost) or np.isnan(monto_sim) or monto_sim <= 0:
        return np.nan, np.nan
    neto = monto_sim / VAT
    if neto == 0:
        return np.nan, np.nan
    margin = (neto - cost) / neto
    return neto, margin


def margin_label(value):
    if pd.isna(value):
        return "—"
    return f"{value:.4f} · {value:.1%}"


def split_mlc_cell(value):
    raw = clean_text(value)
    if not raw:
        return []
    parts = []
    for token in raw.replace("/", "-").split("-"):
        tok = clean_text(token)
        if not tok:
            continue
        digits = "".join(ch for ch in tok if ch.isdigit())
        if not digits:
            continue
        parts.append(f"MLC{digits}")
    # de-duplicate while keeping order
    seen = set()
    out = []
    for p in parts:
        if p not in seen:
            out.append(p)
            seen.add(p)
    return out


def ensure_columns(df, cols):
    for c in cols:
        if c not in df.columns:
            df[c] = np.nan
    return df


# ---------- Cached loaders ----------
@st.cache_data(show_spinner=False)
def load_master_bytes(file_bytes: bytes):
    xls = pd.ExcelFile(io.BytesIO(file_bytes))

    maestra = pd.read_excel(io.BytesIO(file_bytes), sheet_name="MAESTRA de precios")
    maestra = ensure_columns(
        maestra,
        [
            "SKU", "DESCRIPCIÓN", "UBIC", "ÚLTIMO COSTO", "MARGEN LOCAL", "PRECIO NETO", "PRECIO BRUTO",
            "MARGEN MELI 1", " NETO MELI 1", "MONTO EN SIMULACIÓN", "Unnamed: 12", "MLC",
            "PRECIO B2C PUBLICADO ", "FECHA VENCI", "COMENTARIO", "MLC.1", "PRECIO B2C", "FECHA VENCI.1", "COMENTARIO.1"
        ],
    ).copy()
    maestra["SKU_norm"] = maestra["SKU"].apply(clean_sku)
    maestra["DESCRIPCION_norm"] = maestra["DESCRIPCIÓN"].apply(clean_text)
    maestra["FECHA VENCI"] = maestra["FECHA VENCI"].apply(parse_date)
    maestra["FECHA VENCI.1"] = maestra["FECHA VENCI.1"].apply(parse_date)

    bridge = pd.read_excel(io.BytesIO(file_bytes), sheet_name="MLC -SKU")
    bridge = ensure_columns(bridge, ["SKU", "Número de publicación"]).copy()
    bridge["SKU_norm"] = bridge["SKU"].apply(clean_sku)
    bridge["MLC_norm"] = bridge["Número de publicación"].apply(lambda x: split_mlc_cell(x)[0] if split_mlc_cell(x) else "")

    # Relámpago has no true header row
    rel_raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name="Relampago mi pagina", header=None)
    rel_raw = rel_raw.iloc[:, :6].copy()
    while rel_raw.shape[1] < 6:
        rel_raw[rel_raw.shape[1]] = np.nan
    rel_raw.columns = ["SKU", "DESCRIPCIÓN", "PRECIO B2C", "EXTRA", "TIPO", "ESTADO"]
    rel = rel_raw.copy()
    rel["SKU_norm"] = rel["SKU"].apply(clean_sku)
    rel["DESCRIPCIÓN"] = rel["DESCRIPCIÓN"].apply(clean_text)
    rel["PRECIO B2C"] = rel["PRECIO B2C"].apply(to_float)
    rel["TIPO"] = rel["TIPO"].apply(clean_text)
    rel["ESTADO"] = rel["ESTADO"].apply(clean_text)

    return maestra, bridge, rel, rel_raw, xls.sheet_names


@st.cache_data(show_spinner=False)
def load_compras_bytes(file_bytes: bytes):
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0)
    df = ensure_columns(df, ["Fecha", "SKU", "Razón Social", "Precio Un.", "Cantidad", "Concepto / Artículo"]).copy()
    df = df[df["SKU"].notna()].copy()
    df["SKU_norm"] = df["SKU"].apply(clean_sku)
    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce", dayfirst=True)
    df["Precio Un."] = df["Precio Un."].apply(to_float)
    df["Cantidad"] = df["Cantidad"].apply(to_float)
    df["Razón Social"] = df["Razón Social"].apply(clean_text)
    df["Concepto / Artículo"] = df["Concepto / Artículo"].apply(clean_text)
    df = df[df["SKU_norm"] != ""].copy()
    return df


# ---------- Model building ----------
def build_promos_from_maestra(maestra_df: pd.DataFrame) -> pd.DataFrame:
    records = []
    for idx, row in maestra_df.iterrows():
        sku = row.get("SKU_norm", "")
        desc = clean_text(row.get("DESCRIPCIÓN"))

        slots = [
            {
                "slot": 1,
                "mlc_col": "MLC",
                "fallback_mlc_col": "Unnamed: 12",
                "price_col": "PRECIO B2C PUBLICADO ",
                "date_col": "FECHA VENCI",
                "comment_col": "COMENTARIO",
            },
            {
                "slot": 2,
                "mlc_col": "MLC.1",
                "fallback_mlc_col": None,
                "price_col": "PRECIO B2C",
                "date_col": "FECHA VENCI.1",
                "comment_col": "COMENTARIO.1",
            },
        ]

        for slot in slots:
            mlc_tokens = split_mlc_cell(row.get(slot["mlc_col"]))
            if not mlc_tokens and slot["fallback_mlc_col"]:
                mlc_tokens = split_mlc_cell(row.get(slot["fallback_mlc_col"]))
            price = to_float(row.get(slot["price_col"]))
            dt = parse_date(row.get(slot["date_col"]))
            comment = clean_text(row.get(slot["comment_col"]))

            has_payload = bool(mlc_tokens or not np.isnan(price) or not pd.isna(dt) or comment)
            if not has_payload:
                continue

            if not mlc_tokens:
                mlc_tokens = [""]

            for mlc in mlc_tokens:
                records.append(
                    {
                        "row_idx": idx,
                        "slot": slot["slot"],
                        "SKU_norm": sku,
                        "DESCRIPCIÓN": desc,
                        "MLC": mlc,
                        "PRECIO_B2C": price,
                        "FECHA_VENCI": dt,
                        "COMENTARIO": comment,
                        "status": status_from_date(dt),
                        "price_col": slot["price_col"],
                        "date_col": slot["date_col"],
                        "comment_col": slot["comment_col"],
                    }
                )
    promos = pd.DataFrame(records)
    if promos.empty:
        return pd.DataFrame(columns=["row_idx", "slot", "SKU_norm", "DESCRIPCIÓN", "MLC", "PRECIO_B2C", "FECHA_VENCI", "COMENTARIO", "status", "price_col", "date_col", "comment_col"])
    return promos


def build_products(maestra_df: pd.DataFrame, bridge_df: pd.DataFrame, promos_df: pd.DataFrame, rel_df: pd.DataFrame) -> pd.DataFrame:
    bridge_map = bridge_df.groupby("SKU_norm")["MLC_norm"].apply(lambda s: [x for x in s if x]).to_dict()
    promo_map = promos_df.groupby("SKU_norm")["MLC"].apply(lambda s: [x for x in s if x]).to_dict() if not promos_df.empty else {}
    rel_map = rel_df.groupby("SKU_norm").size().to_dict() if not rel_df.empty else {}

    rows = []
    for _, r in maestra_df.iterrows():
        sku = r["SKU_norm"]
        all_mlc = []
        for source in [bridge_map.get(sku, []), promo_map.get(sku, [])]:
            for mlc in source:
                if mlc and mlc not in all_mlc:
                    all_mlc.append(mlc)
        promos_count = int((promos_df["SKU_norm"] == sku).sum()) if not promos_df.empty else 0
        rows.append(
            {
                "SKU_norm": sku,
                "DESCRIPCIÓN": clean_text(r.get("DESCRIPCIÓN")),
                "UBIC": clean_text(r.get("UBIC")),
                "ÚLTIMO COSTO": to_float(r.get("ÚLTIMO COSTO")),
                "PRECIO BRUTO": to_float(r.get("PRECIO BRUTO")),
                "PRECIO NETO": to_float(r.get("PRECIO NETO")),
                "MONTO EN SIMULACIÓN": to_float(r.get("MONTO EN SIMULACIÓN")),
                " NETO MELI 1": to_float(r.get(" NETO MELI 1")),
                "MARGEN LOCAL": to_float(r.get("MARGEN LOCAL")),
                "MARGEN MELI 1": to_float(r.get("MARGEN MELI 1")),
                "all_mlc": all_mlc,
                "promos_count": promos_count,
                "has_relampago": bool(rel_map.get(sku, 0)),
            }
        )
    return pd.DataFrame(rows)


def ensure_model():
    key = (
        st.session_state.get("master_hash", ""),
        st.session_state.get("compras_hash", ""),
        st.session_state.get("maestra_rev", 0),
        st.session_state.get("rel_rev", 0),
    )
    if st.session_state.get("model_key") == key:
        return

    maestra_df = st.session_state["maestra_df"]
    bridge_df = st.session_state["bridge_df"]
    rel_df = st.session_state["rel_df"]
    promos_df = build_promos_from_maestra(maestra_df)
    products_df = build_products(maestra_df, bridge_df, promos_df, rel_df)

    purchase_summary = pd.DataFrame()
    if "compras_df" in st.session_state:
        c = st.session_state["compras_df"].sort_values(["SKU_norm", "Fecha"]).copy()
        if not c.empty:
            summary_rows = []
            for sku, g in c.groupby("SKU_norm"):
                g = g.sort_values("Fecha")
                last = g.iloc[-1]
                prev = g.iloc[-2] if len(g) > 1 else None
                summary_rows.append(
                    {
                        "SKU_norm": sku,
                        "ultima_fecha": last["Fecha"],
                        "ultimo_precio": to_float(last["Precio Un."]),
                        "ultimo_proveedor": clean_text(last["Razón Social"]),
                        "ultima_cantidad": to_float(last["Cantidad"]),
                        "compra_anterior_precio": to_float(prev["Precio Un."]) if prev is not None else np.nan,
                        "variacion_abs": (to_float(last["Precio Un."]) - to_float(prev["Precio Un."])) if prev is not None else np.nan,
                        "variacion_pct": ((to_float(last["Precio Un."]) / to_float(prev["Precio Un."])) - 1) if prev is not None and to_float(prev["Precio Un."]) not in [0, np.nan] else np.nan,
                        "proveedores": ", ".join(sorted({x for x in g["Razón Social"] if x})),
                    }
                )
            purchase_summary = pd.DataFrame(summary_rows)

    st.session_state["model"] = {
        "promos_df": promos_df,
        "products_df": products_df,
        "purchase_summary": purchase_summary,
    }
    st.session_state["model_key"] = key


# ---------- Workbook export ----------
def update_sheet_from_dataframe(ws, df: pd.DataFrame, include_header=True, header_row=1):
    start_row = header_row + (1 if include_header else 0)
    # clear existing values below header
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(start_row, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(r, c).value = None

    if include_header:
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(header_row, col_idx).value = col_name

    for r_idx, (_, row) in enumerate(df.iterrows(), start=start_row):
        for c_idx, val in enumerate(row.tolist(), start=1):
            cell = ws.cell(r_idx, c_idx)
            if pd.isna(val):
                cell.value = None
            elif isinstance(val, pd.Timestamp):
                cell.value = val.to_pydatetime()
                cell.number_format = "DD/MM/YYYY"
            else:
                cell.value = val


def build_download_bytes(master_bytes: bytes, maestra_df: pd.DataFrame, rel_raw_export: pd.DataFrame) -> bytes:
    wb = load_workbook(io.BytesIO(master_bytes))
    # maestra without helper cols
    maestra_export = maestra_df.drop(columns=[c for c in ["SKU_norm", "DESCRIPCION_norm"] if c in maestra_df.columns])
    ws = wb["MAESTRA de precios"]
    update_sheet_from_dataframe(ws, maestra_export, include_header=True, header_row=1)

    ws_rel = wb["Relampago mi pagina"]
    update_sheet_from_dataframe(ws_rel, rel_raw_export, include_header=False, header_row=1)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ---------- State init from uploads ----------
def init_from_uploads(master_file, compras_file):
    if master_file is None:
        return
    current_hash = file_hash(master_file)
    if st.session_state.get("master_hash") != current_hash:
        maestra, bridge, rel, rel_raw, sheet_names = load_master_bytes(master_file.getvalue())
        st.session_state["master_hash"] = current_hash
        st.session_state["master_bytes"] = master_file.getvalue()
        st.session_state["maestra_df"] = maestra.copy()
        st.session_state["bridge_df"] = bridge.copy()
        st.session_state["rel_df"] = rel.copy()
        st.session_state["rel_raw_df"] = rel_raw.copy()
        st.session_state["sheet_names"] = sheet_names
        st.session_state["maestra_rev"] = 0
        st.session_state["rel_rev"] = 0
        st.session_state["download_cache_key"] = None

    if compras_file is not None:
        c_hash = file_hash(compras_file)
        if st.session_state.get("compras_hash") != c_hash:
            st.session_state["compras_hash"] = c_hash
            st.session_state["compras_df"] = load_compras_bytes(compras_file.getvalue()).copy()
    else:
        st.session_state.pop("compras_hash", None)
        st.session_state.pop("compras_df", None)


# ---------- Sidebar ----------
st.sidebar.header("Archivos")
master_file = st.sidebar.file_uploader("Maestra saneada", type=["xlsx"], key="master_upload")
compras_file = st.sidebar.file_uploader("Compras (opcional)", type=["xlsx"], key="compras_upload")
init_from_uploads(master_file, compras_file)

if "maestra_df" not in st.session_state:
    st.title("Aurora · Precios y Promos")
    st.info("Sube la **MAESTRA PRECIOS Y PROMOS (4).xlsx** para comenzar.")
    st.stop()

ensure_model()
maestra_df = st.session_state["maestra_df"]
bridge_df = st.session_state["bridge_df"]
rel_df = st.session_state["rel_df"]
rel_raw_df = st.session_state["rel_raw_df"]
model = st.session_state["model"]
promos_df = model["promos_df"]
products_df = model["products_df"]
purchase_summary = model["purchase_summary"]
compras_df = st.session_state.get("compras_df", pd.DataFrame())

st.title("Aurora · Precios y Promos")
st.caption("Modelo nuevo, desde cero. Solo usa MAESTRA de precios, MLC -SKU, Relámpago mi página y Compras.")


# ---------- Dialogs ----------
@st.dialog("Editar promo", width="large")
def promo_dialog(promo_idx: int):
    promo = promos_df.loc[promo_idx]
    st.markdown(f"**SKU:** {promo['SKU_norm']}  ")
    st.markdown(f"**Descripción:** {promo['DESCRIPCIÓN']}")
    mlc_show = promo['MLC'] or "Sin MLC"
    st.markdown(f"**MLC:** {mlc_show}")

    new_date = st.date_input(
        "Fecha de vencimiento",
        value=promo["FECHA_VENCI"].date() if not pd.isna(promo["FECHA_VENCI"]) else TODAY,
        format="DD/MM/YYYY",
    )
    with st.expander("Precio B2C y comentario", expanded=False):
        new_price = st.number_input("Precio B2C publicado", min_value=0.0, value=float(promo["PRECIO_B2C"]) if not np.isnan(promo["PRECIO_B2C"]) else 0.0, step=100.0)
        new_comment = st.text_input("Comentario", value=promo["COMENTARIO"])

    c1, c2 = st.columns(2)
    if c1.button("Guardar", type="primary", use_container_width=True):
        row_idx = int(promo["row_idx"])
        st.session_state["maestra_df"].at[row_idx, promo["price_col"]] = new_price if new_price > 0 else np.nan
        st.session_state["maestra_df"].at[row_idx, promo["date_col"]] = pd.Timestamp(new_date)
        st.session_state["maestra_df"].at[row_idx, promo["comment_col"]] = new_comment.strip() or np.nan
        st.session_state["maestra_rev"] += 1
        st.session_state["download_cache_key"] = None
        st.rerun()
    if c2.button("Cancelar", use_container_width=True):
        st.rerun()


# ---------- Tabs ----------
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "Cockpit por producto",
    "Operador de promos",
    "Relámpago",
    "Alta de producto",
    "Descarga",
])

# ---------- Cockpit ----------
with tab1:
    search_df = products_df.copy()
    search_df["label"] = search_df.apply(lambda r: f"{r['SKU_norm']} · {r['DESCRIPCIÓN']}", axis=1)
    selected_label = st.selectbox("Buscar producto", options=search_df["label"].tolist(), index=0)
    selected = search_df[search_df["label"] == selected_label].iloc[0]
    sku = selected["SKU_norm"]
    selected_maestra = maestra_df[maestra_df["SKU_norm"] == sku].iloc[0]
    product_promos = promos_df[promos_df["SKU_norm"] == sku].copy()
    rel_product = rel_df[rel_df["SKU_norm"] == sku].copy()
    product_purchases = compras_df[compras_df["SKU_norm"] == sku].sort_values("Fecha") if not compras_df.empty else pd.DataFrame()
    purchase_row = purchase_summary[purchase_summary["SKU_norm"] == sku]

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Último costo", "—" if pd.isna(selected["ÚLTIMO COSTO"]) else f"${selected['ÚLTIMO COSTO']:,.0f}")
    c2.metric("Precio bruto tienda", "—" if pd.isna(selected["PRECIO BRUTO"]) else f"${selected['PRECIO BRUTO']:,.0f}")
    c3.metric("Margen local", margin_label(selected["MARGEN LOCAL"]))
    c4.metric("Margen Meli 1", margin_label(selected["MARGEN MELI 1"]))

    c5, c6, c7 = st.columns([1.3, 1.3, 2])
    with c5:
        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        st.markdown("**Precio y rentabilidad**")
        st.write(f"**Precio neto tienda:** {'—' if pd.isna(selected['PRECIO NETO']) else f'$ {selected['PRECIO NETO']:,.0f}'}")
        st.write(f"**Monto en simulación:** {'—' if pd.isna(selected['MONTO EN SIMULACIÓN']) else f'$ {selected['MONTO EN SIMULACIÓN']:,.0f}'}")
        st.write(f"**Neto Meli 1:** {'—' if pd.isna(selected[' NETO MELI 1']) else f'$ {selected[' NETO MELI 1']:,.0f}'}")
        st.write(f"**Ubicación:** {clean_text(selected['UBIC']) or '—'}")
        st.markdown('</div>', unsafe_allow_html=True)
    with c6:
        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        st.markdown("**Publicaciones MLC asociadas**")
        if selected["all_mlc"]:
            st.markdown("".join([f'<span class="chip">{m}</span>' for m in selected["all_mlc"]]), unsafe_allow_html=True)
        else:
            st.caption("Sin MLC asociado")
        st.markdown('</div>', unsafe_allow_html=True)
    with c7:
        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        st.markdown("**Promos activas en maestra**")
        if product_promos.empty:
            st.info("No tiene promo cargada en la maestra.")
        else:
            view = product_promos[["MLC", "PRECIO_B2C", "FECHA_VENCI", "COMENTARIO", "status"]].copy()
            view.columns = ["MLC", "Precio B2C", "Fecha", "Comentario", "Estado"]
            view["Fecha"] = view["Fecha"].apply(fmt_date)
            st.dataframe(view, use_container_width=True, hide_index=True)
        st.markdown('</div>', unsafe_allow_html=True)

    col_rel, col_buy = st.columns([1, 1.4])
    with col_rel:
        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        st.markdown("**Relámpago mi página**")
        if rel_product.empty:
            st.caption("No está en relámpago.")
        else:
            rel_view = rel_product[["SKU_norm", "DESCRIPCIÓN", "PRECIO B2C", "TIPO", "ESTADO"]].copy()
            rel_view.columns = ["SKU", "Descripción", "Precio B2C", "Tipo", "Estado"]
            st.dataframe(rel_view, use_container_width=True, hide_index=True)
        st.markdown('</div>', unsafe_allow_html=True)
    with col_buy:
        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        st.markdown("**Compras**")
        if purchase_row.empty:
            st.caption("No hay compras encontradas para este SKU.")
        else:
            pr = purchase_row.iloc[0]
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Última compra", fmt_date(pr["ultima_fecha"]))
            m2.metric("Último precio", f"${pr['ultimo_precio']:,.0f}" if not pd.isna(pr["ultimo_precio"]) else "—")
            m3.metric("Proveedor", pr["ultimo_proveedor"] or "—")
            m4.metric("Variación", "—" if pd.isna(pr["variacion_pct"]) else f"{pr['variacion_pct']:.1%}")
            if not product_purchases.empty:
                hist = product_purchases[["Fecha", "Precio Un.", "Cantidad", "Razón Social", "Concepto / Artículo"]].copy()
                hist.columns = ["Fecha", "Precio Un.", "Cantidad", "Proveedor", "Artículo"]
                hist["Fecha"] = hist["Fecha"].apply(fmt_date)
                st.dataframe(hist.sort_values("Fecha", ascending=False), use_container_width=True, hide_index=True, height=240)
        st.markdown('</div>', unsafe_allow_html=True)

# ---------- Operador ----------
with tab2:
    st.subheader("Operador de promos")
    left, right = st.columns([1.25, 1])
    with left:
        search = st.text_input("Buscar por SKU, descripción o MLC")
    with right:
        status_filter = st.selectbox(
            "Filtro",
            ["Todas", "Vence hoy", "Vence mañana", "Vence en 3 días", "Próx. 7 días", "Vencida", "Sin fecha", "Sin precio B2C"],
        )

    filtered = promos_df.copy()
    if search:
        q = search.lower().strip()
        filtered = filtered[
            filtered["SKU_norm"].str.lower().str.contains(q)
            | filtered["DESCRIPCIÓN"].str.lower().str.contains(q)
            | filtered["MLC"].str.lower().str.contains(q)
        ]
    if status_filter == "Sin precio B2C":
        filtered = filtered[filtered["PRECIO_B2C"].isna()]
    elif status_filter != "Todas":
        filtered = filtered[filtered["status"] == status_filter]

    with st.expander("Actualización masiva de fecha", expanded=False):
        new_mass_date = st.date_input("Nueva fecha para las promos filtradas", value=TODAY, format="DD/MM/YYYY", key="mass_date")
        if st.button(f"Aplicar fecha a {len(filtered)} promo(s)", type="primary"):
            for _, promo in filtered.iterrows():
                row_idx = int(promo["row_idx"])
                st.session_state["maestra_df"].at[row_idx, promo["date_col"]] = pd.Timestamp(new_mass_date)
            st.session_state["maestra_rev"] += 1
            st.session_state["download_cache_key"] = None
            st.rerun()

    counts = filtered["status"].value_counts().to_dict() if not filtered.empty else {}
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Filtradas", len(filtered))
    c2.metric("Vence hoy", counts.get("Vence hoy", 0))
    c3.metric("Mañana", counts.get("Vence mañana", 0))
    c4.metric("En 3 días", counts.get("Vence en 3 días", 0))

    if filtered.empty:
        st.info("No hay promos para este filtro.")
    else:
        display = filtered.sort_values(["FECHA_VENCI", "SKU_norm"], na_position="last")
        cols = st.columns(4)
        for i, (pidx, promo) in enumerate(display.iterrows()):
            with cols[i % 4]:
                label = f"{promo['SKU_norm']}\n{promo['DESCRIPCIÓN'][:44]}\n{promo['MLC'] or 'Sin MLC'}\n{fmt_date(promo['FECHA_VENCI'])}"
                if st.button(label, key=f"promo_card_{pidx}", use_container_width=True, type="secondary"):
                    promo_dialog(pidx)

# ---------- Relámpago ----------
with tab3:
    st.subheader("Relámpago mi página")
    st.caption("Lista simple para agregar, quitar o modificar relámpagos.")
    rel_editor = rel_df[["SKU", "DESCRIPCIÓN", "PRECIO B2C", "TIPO", "ESTADO"]].copy()
    edited_rel = st.data_editor(
        rel_editor,
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        column_config={
            "SKU": st.column_config.TextColumn("SKU"),
            "DESCRIPCIÓN": st.column_config.TextColumn("Descripción", width="large"),
            "PRECIO B2C": st.column_config.NumberColumn("Precio B2C", step=100),
            "TIPO": st.column_config.TextColumn("Tipo"),
            "ESTADO": st.column_config.TextColumn("Estado"),
        },
        key="rel_editor",
    )
    if st.button("Guardar relámpago", type="primary"):
        new_rel = edited_rel.copy()
        new_rel["SKU_norm"] = new_rel["SKU"].apply(clean_sku)
        new_rel["DESCRIPCIÓN"] = new_rel["DESCRIPCIÓN"].apply(clean_text)
        new_rel["TIPO"] = new_rel["TIPO"].apply(clean_text)
        new_rel["ESTADO"] = new_rel["ESTADO"].apply(clean_text)
        st.session_state["rel_df"] = new_rel
        # raw export is headerless with original 6 cols
        raw_export = new_rel[["SKU", "DESCRIPCIÓN", "PRECIO B2C"]].copy()
        raw_export[3] = np.nan
        raw_export[4] = new_rel["TIPO"]
        raw_export[5] = new_rel["ESTADO"]
        st.session_state["rel_raw_df"] = raw_export
        st.session_state["rel_rev"] += 1
        st.session_state["download_cache_key"] = None
        st.rerun()

# ---------- Alta ----------
with tab4:
    st.subheader("Alta de producto")
    sku_new = st.text_input("SKU nuevo")
    desc_new = st.text_input("Descripción")
    ubic_new = st.text_input("Ubicación")

    a1, a2, a3 = st.columns(3)
    costo_new = a1.number_input("Último costo", min_value=0.0, step=100.0)
    bruto_new = a2.number_input("Precio bruto en tienda", min_value=0.0, step=100.0)
    sim_new = a3.number_input("Monto en simulación", min_value=0.0, step=100.0)

    neto_local, m_local = margin_local(costo_new, bruto_new)
    neto_m1, m_m1 = margin_meli1(costo_new, sim_new)
    b1, b2, b3, b4 = st.columns(4)
    b1.metric("Precio neto tienda", "—" if np.isnan(neto_local) else f"${neto_local:,.0f}")
    b2.metric("Margen local", margin_label(m_local))
    b3.metric("Neto Meli 1", "—" if np.isnan(neto_m1) else f"${neto_m1:,.0f}")
    b4.metric("Margen Meli 1", margin_label(m_m1))

    st.markdown("**Promo base en maestra**")
    p1, p2, p3 = st.columns(3)
    mlc1 = p1.text_input("MLC slot 1")
    b2c1 = p2.number_input("Precio B2C publicado", min_value=0.0, step=100.0)
    fecha1 = p3.date_input("Fecha venci slot 1", value=TODAY, format="DD/MM/YYYY")
    comentario1 = st.text_input("Comentario slot 1")

    q1, q2, q3 = st.columns(3)
    mlc2 = q1.text_input("MLC slot 2")
    b2c2 = q2.number_input("Precio B2C slot 2", min_value=0.0, step=100.0)
    fecha2 = q3.date_input("Fecha venci slot 2", value=TODAY, format="DD/MM/YYYY")
    comentario2 = st.text_input("Comentario slot 2")

    add_rel = st.checkbox("Agregar también a relámpago")
    rel_tipo = st.text_input("Tipo relámpago", value="LIQUIDACION") if add_rel else ""
    rel_estado = st.text_input("Estado relámpago", value="") if add_rel else ""

    if st.button("Crear producto", type="primary"):
        sku_norm = clean_sku(sku_new)
        if not sku_norm:
            st.error("Ingresa un SKU válido.")
        elif sku_norm in set(maestra_df["SKU_norm"]):
            st.error("Ese SKU ya existe en la maestra.")
        else:
            new_row = {col: np.nan for col in maestra_df.columns}
            new_row.update(
                {
                    "SKU": sku_norm,
                    "DESCRIPCIÓN": desc_new,
                    "UBIC": ubic_new,
                    "ÚLTIMO COSTO": costo_new if costo_new > 0 else np.nan,
                    "PRECIO BRUTO": bruto_new if bruto_new > 0 else np.nan,
                    "PRECIO NETO": neto_local if not np.isnan(neto_local) else np.nan,
                    "MARGEN LOCAL": m_local if not np.isnan(m_local) else np.nan,
                    "MONTO EN SIMULACIÓN": sim_new if sim_new > 0 else np.nan,
                    " NETO MELI 1": neto_m1 if not np.isnan(neto_m1) else np.nan,
                    "MARGEN MELI 1": m_m1 if not np.isnan(m_m1) else np.nan,
                    "MLC": mlc1 or np.nan,
                    "PRECIO B2C PUBLICADO ": b2c1 if b2c1 > 0 else np.nan,
                    "FECHA VENCI": pd.Timestamp(fecha1) if b2c1 > 0 or mlc1 else np.nan,
                    "COMENTARIO": comentario1 or np.nan,
                    "MLC.1": mlc2 or np.nan,
                    "PRECIO B2C": b2c2 if b2c2 > 0 else np.nan,
                    "FECHA VENCI.1": pd.Timestamp(fecha2) if b2c2 > 0 or mlc2 else np.nan,
                    "COMENTARIO.1": comentario2 or np.nan,
                    "SKU_norm": sku_norm,
                    "DESCRIPCION_norm": desc_new,
                }
            )
            st.session_state["maestra_df"] = pd.concat([st.session_state["maestra_df"], pd.DataFrame([new_row])], ignore_index=True)
            # also add bridge mlcs
            bridge_adds = []
            for mlc in split_mlc_cell(mlc1) + split_mlc_cell(mlc2):
                bridge_adds.append({"SKU": sku_norm, "Número de publicación": mlc, "SKU_norm": sku_norm, "MLC_norm": mlc})
            if bridge_adds:
                st.session_state["bridge_df"] = pd.concat([st.session_state["bridge_df"], pd.DataFrame(bridge_adds)], ignore_index=True)
            if add_rel and b2c1 > 0:
                new_rel = {"SKU": sku_norm, "DESCRIPCIÓN": desc_new, "PRECIO B2C": b2c1, "TIPO": rel_tipo, "ESTADO": rel_estado, "SKU_norm": sku_norm}
                st.session_state["rel_df"] = pd.concat([st.session_state["rel_df"], pd.DataFrame([new_rel])], ignore_index=True)
                raw = st.session_state["rel_raw_df"].copy()
                raw = pd.concat([raw, pd.DataFrame([[sku_norm, desc_new, b2c1, np.nan, rel_tipo, rel_estado]])], ignore_index=True)
                st.session_state["rel_raw_df"] = raw
                st.session_state["rel_rev"] += 1
            st.session_state["maestra_rev"] += 1
            st.session_state["download_cache_key"] = None
            st.success("Producto creado en memoria. Descárgalo en la pestaña de descarga.")
            st.rerun()

# ---------- Download ----------
with tab5:
    st.subheader("Descarga")
    st.caption("La app trabaja en memoria. Aquí descargas la maestra actualizada.")
    cache_key = (
        st.session_state.get("master_hash", ""),
        st.session_state.get("maestra_rev", 0),
        st.session_state.get("rel_rev", 0),
    )
    if st.session_state.get("download_cache_key") != cache_key:
        st.session_state["download_bytes"] = build_download_bytes(
            st.session_state["master_bytes"],
            st.session_state["maestra_df"],
            st.session_state["rel_raw_df"],
        )
        st.session_state["download_cache_key"] = cache_key
    st.download_button(
        "Descargar Excel actualizado",
        data=st.session_state["download_bytes"],
        file_name="MAESTRA_PRECIOS_PROMOS_ACTUALIZADA.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )
